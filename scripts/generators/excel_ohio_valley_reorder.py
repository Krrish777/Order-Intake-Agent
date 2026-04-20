"""Generate Ohio Valley Industrial's reorder sheet — Phase 1.3 ambiguity exemplar.

Ohio Valley Industrial runs a legacy JD Edwards ERP in Cincinnati that exports
weekly reorder sheets as .xlsx. Per their customer master, "headers vary by
export template; they've sent three different layouts since 2023." This is the
third-version (current) layout — machine-generated, sparse, and missing several
fields that would normally anchor an extraction.

This file's edge case is **ambiguity via omission**, not typos. The agent must
recognize the missing fields and route for human review rather than guess:

1.  **No explicit PO number.** The filename is the only reference.
    (`ohio_valley_reorder_march_wk3.xlsx` — not even a real date, just a week
    tag that recycles.)
2.  **No prices column.** JDE stock-status exports list what the planner needs
    to order, not what Grafton-Reese charges. Agent must look up contract prices.
3.  **"Need By: ASAP" instead of a structured date.** Human-meaningful,
    machine-unparseable. Forces escalation or default lead-time logic.
4.  **Header row at row 4** (after a 3-row branding block) — not row 1, not
    row 13 like Hagan. Positional header heuristics fail without a scan.
5.  **Machine-export aesthetic** — all-caps headers, `END OF REPORT` footer,
    no formulas, no styling flourishes. Reads as generated-not-authored.
6.  **Sparse line count (5).** Ohio Valley reorders are typically narrow: they
    run MRP weekly and only surface what fell below the reorder point.
7.  **No sku_aliases.** Canonical Grafton-Reese SKUs throughout.

Run with: ``uv run python -m scripts.generators.excel_ohio_valley_reorder``
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side


@dataclass(frozen=True)
class LineItem:
    item_nbr: str
    description: str
    qty_req: int
    uom: str


# Five items that would realistically drop below reorder point at a Cincinnati
# general-industrial distributor: a Grade 8 structural fastener, a nyloc commonly
# consumed, two pneumatic items for shop-floor air lines, and an R2 hose reel.
# Quantities reflect MRP-driven replenishment (above MOQ, clean pack multiples
# where the planner rounded to box size, non-round where they did not).
LINES: tuple[LineItem, ...] = (
    LineItem("FST-HCS-038-16-100-G8YZ", "HCS 3/8-16 X 1 GR8 YELLOW ZINC",          500, "EA"),
    LineItem("FST-NYL-038-16-S18",      "NYLOC 3/8-16 18-8 SS",                    300, "EA"),
    LineItem("PNM-PCS-025-025N-BR",     "PUSH-CONNECT STR 1/4T X 1/4NPT BRASS",     50, "EA"),
    LineItem("PNM-QCI-025N-STL",        "QC COUPLER IND 1/4 NPT F STL",             40, "EA"),
    LineItem("HYD-HSE-R2-06",           "HOSE 100R2AT -6 3/8 ID 2-WIRE",           100, "FT"),
)


def _thin_border() -> Border:
    side = Side(style="thin", color="BFBFBF")
    return Border(left=side, right=side, top=side, bottom=side)


def build_workbook() -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "RPT_MRO_V3"  # JD Edwards report ID — the "v3" of three layouts

    # Sparse 4-column layout: ITEM NBR | DESCRIPTION | QTY REQ | UOM
    # Deliberately no PRICE column — the extractor must not fabricate one.
    widths = {"A": 26, "B": 42, "C": 10, "D": 7}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # --- Rows 1-3: compact branding / metadata block ---
    # Mono-style ERP export header: minimal styling, small font, left-aligned.
    banner_font = Font(name="Consolas", size=11, bold=True)
    meta_font = Font(name="Consolas", size=9)

    ws["A1"] = "OHIO VALLEY INDUSTRIAL CO."
    ws["A1"].font = banner_font
    ws.merge_cells("A1:D1")

    # Report metadata — note no PO number field
    ws["A2"] = "RPT ID: RPT_MRO_V3    RUN: 03/16/26 04:12   BRANCH: CIN-01"
    ws["A2"].font = meta_font
    ws.merge_cells("A2:D2")

    # Ship-to + need-by compressed onto one line — "ASAP" as a literal string
    ws["A3"] = "SHIP TO: OVI-CIN   NEED BY: ASAP   TERMS: NET 45   VENDOR: GRAFTON-REESE"
    ws["A3"].font = meta_font
    ws.merge_cells("A3:D3")

    # --- Row 4: column headers ---
    header_row = 4
    headers = ["ITEM NBR", "DESCRIPTION", "QTY REQ", "UOM"]
    header_font = Font(name="Consolas", size=10, bold=True)
    for col, text in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=text)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="left" if col <= 2 else "center", vertical="center")
        cell.border = _thin_border()
    ws.row_dimensions[header_row].height = 16

    # --- Rows 5+: data ---
    body_font = Font(name="Consolas", size=10)
    qty_fmt = "#,##0"
    for i, line in enumerate(LINES, start=1):
        r = header_row + i
        ws.cell(row=r, column=1, value=line.item_nbr).alignment = Alignment(horizontal="left")
        ws.cell(row=r, column=2, value=line.description).alignment = Alignment(horizontal="left")
        ws.cell(row=r, column=3, value=line.qty_req).number_format = qty_fmt
        ws.cell(row=r, column=3).alignment = Alignment(horizontal="right")
        ws.cell(row=r, column=4, value=line.uom).alignment = Alignment(horizontal="center")

        for col in range(1, 5):
            ws.cell(row=r, column=col).font = body_font

    last_line_row = header_row + len(LINES)

    # --- Machine-generated footer ---
    footer_row = last_line_row + 2
    ws.cell(row=footer_row, column=1,
            value=f"END OF REPORT — {len(LINES)} LINES").font = Font(
        name="Consolas", size=9, italic=True, color="595959"
    )
    ws.merge_cells(start_row=footer_row, start_column=1, end_row=footer_row, end_column=4)

    # No totals block (no prices to sum), no signature block (machine export),
    # no special instructions. The sparseness is the edge case.

    return wb


def main() -> Path:
    out = Path(__file__).resolve().parents[2] / "data" / "excel" / "ohio_valley_reorder_march_wk3.xlsx"
    out.parent.mkdir(parents=True, exist_ok=True)
    wb = build_workbook()
    wb.save(out)
    return out


if __name__ == "__main__":
    path = main()
    print(f"wrote: {path}")
