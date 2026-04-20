"""Generate Hagan BPT's reorder spreadsheet — Phase 1.2 typo-variant exemplar.

Hagan Bearing & Power Transmission is a regional-chain distributor in Pittsburgh's
Strip District with a small branch in Wheeling WV. Unlike GLFP, they have no EDI and
no corporate PO template — their "order sheet" is a homemade Excel file that was
created years ago and has carried the same typos through dozens of versions. The
customer notes flag that ship-to varies (Pittsburgh vs Wheeling) so that must be
verified on every order.

This is the typo-and-label-variation exemplar. Intentional realism anchors:

1.  **Template-frozen header typos.** `Qty Ordred`, `Due Dte`, `Desciption` — not
    fresh typos, but ones baked into the template long ago. A procurement manager
    would recognize this pattern instantly (nobody ever fixes these in a working
    template because the macros/formulas reference them).
2.  **Column reorder.** UOM appears BEFORE quantity rather than after, which breaks
    any positional extraction that assumes "Qty then UOM" ordering.
3.  **Inconsistent label casing.** `PO No:` but `po date:`, `Ship To:` but
    `payment terms`. Classic copy-paste template drift.
4.  **Ship-to = Wheeling WV branch** (not Pittsburgh main). Hagan's master record
    warns ship-to varies; using the secondary exercises that validation.
5.  **No contract pricing.** Hagan has no `sku_aliases` and no negotiated prices,
    so line-item costs are Grafton-Reese list price to the cent.
6.  **Mixed hydraulic + fasteners.** Bearing shops buy both — fasteners for
    mountings and JIC/ORB fittings for the power-transmission side.
7.  **PO number style.** `H-26189` — short internal counter, no year prefix. Small
    regional shops rarely use structured PO numbering.

Run with: ``uv run python -m scripts.generators.excel_hagan_reorder``
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter


@dataclass(frozen=True)
class LineItem:
    part_num: str
    description: str
    uom: str
    qty: int
    unit_cost: float


# List prices (from data/masters/products.json) — no Hagan contract discount.
# Qty chosen to look like real mid-order replenishment: not round, above min
# order qty, and believable multiples of pack_size where a bearing shop would
# buy partial boxes (HCS 150 = 3 boxes of 50; flat washers 250 = 2.5 BX of 100).
LINES: tuple[LineItem, ...] = (
    LineItem("FST-HCS-050-13-200-G5Z", "Hex Cap Screw 1/2-13 x 2 GR5 Zinc", "EA", 150, 0.34),
    LineItem("FST-HXN-050-13-G5Z",     "Hex Nut 1/2-13 GR5 Zinc",            "EA", 200, 0.11),
    LineItem("FST-FWS-050-SAE-S18",    "Flat Washer 1/2 SAE 18-8 SS",        "EA", 250, 0.09),
    LineItem("HYD-MJS-06-STL",         "JIC Male Straight -6 Steel",         "EA",  24, 3.42),
    LineItem("HYD-MJE-06-STL",         "JIC Male Elbow 90 -6 Steel",         "EA",  12, 8.75),
    LineItem("HYD-ORB-06-STL",         "ORB Male Straight -6 Steel",         "EA",  15, 5.62),
    LineItem("HYD-MJN-06-04-STL",      "JIC M x NPT M -6 x 1/4 Steel",       "EA",  20, 4.18),
)


def _thin_border() -> Border:
    side = Side(style="thin", color="A6A6A6")
    return Border(left=side, right=side, top=side, bottom=side)


def build_workbook() -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Reorder"

    # Column layout — note UOM (col D) precedes Qty Ordred (col E). That swap is
    # the edge case: positional parsers break if they assume qty-then-uom.
    # Columns: A=Line, B=Part Num, C=Desciption, D=UOM, E=Qty Ordred, F=Unit Cost, G=Ext Total
    widths = {"A": 6, "B": 26, "C": 42, "D": 7, "E": 12, "F": 11, "G": 12}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # Letterhead — plain bold Arial, no corporate color scheme (small-shop tell)
    ws["A1"] = "HAGAN BEARING & POWER TRANSMISSION, INC."
    ws["A1"].font = Font(name="Arial", size=14, bold=True)
    ws.merge_cells("A1:G1")
    ws["A1"].alignment = Alignment(horizontal="center")

    ws["A2"] = "2100 Smallman Street  Pittsburgh PA 15222   ph (412) 281-4906"
    ws["A2"].font = Font(name="Arial", size=9)
    ws.merge_cells("A2:G2")
    ws["A2"].alignment = Alignment(horizontal="center")

    ws["A3"] = "Hagan BPT — Strip District & Wheeling WV branch"
    ws["A3"].font = Font(name="Arial", size=9, italic=True, color="595959")
    ws.merge_cells("A3:G3")
    ws["A3"].alignment = Alignment(horizontal="center")

    # Banner — note "Purchase Order / Reorder" mixed label (template has both)
    ws["A5"] = "Purchase Order / Reorder"
    ws["A5"].font = Font(name="Arial", size=12, bold=True)
    ws.merge_cells("A5:G5")
    ws["A5"].alignment = Alignment(horizontal="center")

    # Metadata block — inconsistent capitalization on labels is the template drift tell.
    # "PO No:" / "po date:" / "Ship To:" / "payment terms" — mixed on purpose.
    bold = Font(name="Arial", size=10, bold=True)
    norm = Font(name="Arial", size=10)

    meta = [
        ("PO No:",          "H-26189",                       "Buyer:",          "Mike Hagan Jr."),
        ("po date:",        "04/09/2026",                    "Email:",          "mike.jr@haganbpt.com"),
        ("Due Dte:",        "04/23/2026",                    "Phone:",          "(412) 281-4906"),
        ("Ship To:",        "Hagan BPT — Wheeling Branch",   "payment terms",   "Net 30"),
        ("",                "401 Main Street, Wheeling WV 26003", "Vendor:",    "Grafton-Reese MRO"),
    ]
    for i, (lab_l, val_l, lab_r, val_r) in enumerate(meta, start=7):
        ws.cell(row=i, column=1, value=lab_l).font = bold
        ws.cell(row=i, column=2, value=val_l).font = norm
        ws.cell(row=i, column=5, value=lab_r).font = bold
        ws.cell(row=i, column=6, value=val_r).font = norm
        ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=4)
        ws.merge_cells(start_row=i, start_column=6, end_row=i, end_column=7)

    # Header row — these typos are intentional and template-frozen
    header_row = 13
    headers = ["Line", "Part Num", "Desciption", "UOM", "Qty Ordred", "Unit Cost", "Ext Total"]
    header_font = Font(name="Arial", size=10, bold=True)
    for col, text in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=text)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _thin_border()
    ws.row_dimensions[header_row].height = 18

    # Line item rows
    body_font = Font(name="Arial", size=10)
    money_fmt = "$#,##0.00"
    qty_fmt = "#,##0"
    for i, line in enumerate(LINES, start=1):
        r = header_row + i
        ws.cell(row=r, column=1, value=i).alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=2, value=line.part_num)
        ws.cell(row=r, column=3, value=line.description)
        ws.cell(row=r, column=3).alignment = Alignment(wrap_text=True, vertical="center")
        ws.cell(row=r, column=4, value=line.uom).alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=5, value=line.qty).number_format = qty_fmt
        ws.cell(row=r, column=5).alignment = Alignment(horizontal="right")
        ws.cell(row=r, column=6, value=line.unit_cost).number_format = money_fmt
        # Ext total = qty * unit cost. Note the qty column is E (not D as in GLFP).
        ws.cell(row=r, column=7, value=f"=E{r}*F{r}").number_format = money_fmt

        for col in range(1, 8):
            ws.cell(row=r, column=col).font = body_font
            ws.cell(row=r, column=col).border = _thin_border()

    last_line_row = header_row + len(LINES)

    # Totals block
    total_row = last_line_row + 2
    ws.cell(row=total_row, column=6, value="Subtotal").font = bold
    ws.cell(row=total_row, column=6).alignment = Alignment(horizontal="right")
    ws.cell(row=total_row, column=7,
            value=f"=SUM(G{header_row + 1}:G{last_line_row})").number_format = money_fmt
    ws.cell(row=total_row, column=7).font = bold

    ws.cell(row=total_row + 1, column=6, value="TOTAL").font = Font(name="Arial", size=11, bold=True)
    ws.cell(row=total_row + 1, column=6).alignment = Alignment(horizontal="right")
    ws.cell(row=total_row + 1, column=7, value=f"=G{total_row}").number_format = money_fmt
    ws.cell(row=total_row + 1, column=7).font = Font(name="Arial", size=11, bold=True)
    ws.cell(row=total_row + 1, column=7).border = Border(
        top=Side(style="thin"), bottom=Side(style="double")
    )

    # Notes — informal small-shop voice
    notes_row = total_row + 3
    ws.cell(row=notes_row, column=1, value="Notes:").font = bold
    ws.cell(row=notes_row + 1, column=1, value=(
        "Ship to Wheeling branch this time — Mike picking up Thurs. "
        "Call if JIC items backordered more than a week."
    ))
    ws.cell(row=notes_row + 1, column=1).alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=notes_row + 1, start_column=1, end_row=notes_row + 2, end_column=7)
    ws.row_dimensions[notes_row + 1].height = 28

    # Sign-off — single line, no legal boilerplate (small-shop tell)
    sig_row = notes_row + 4
    ws.cell(row=sig_row, column=1, value="Authorized by: M. Hagan Jr.")
    ws.cell(row=sig_row, column=1).font = Font(name="Arial", size=9, italic=True)

    return wb


def main() -> Path:
    out = Path(__file__).resolve().parents[2] / "data" / "excel" / "hagan_reorder_2026-04-09.xlsx"
    out.parent.mkdir(parents=True, exist_ok=True)
    wb = build_workbook()
    wb.save(out)
    return out


if __name__ == "__main__":
    path = main()
    print(f"wrote: {path}")
