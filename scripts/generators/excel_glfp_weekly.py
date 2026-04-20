"""Generate GLFP's weekly Tuesday reorder spreadsheet.

GLFP (Great Lakes Fluid Power Group) is a large hydraulics-heavy distributor that
routes weekly reorders through a standardized corporate Excel template. Per the
customer master, they use prefixed SKU aliases (GLFP-XXXXX) that must be resolved
to canonical Grafton-Reese SKUs downstream.

This is the Phase 1.1 clean exemplar. It is deliberately well-formed — no typos,
clean header row, formulaic extended-price column — so the extractor can be
tuned on a known-good baseline before hitting the messier variants.
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


@dataclass(frozen=True)
class LineItem:
    glfp_part: str
    description: str
    qty: float
    uom: str
    unit_price: float


# GLFP contract pricing (12–15% off Grafton-Reese list). Values chosen to reflect
# real contract variance rather than a flat percentage — closer to how a real
# negotiated price sheet would look.
LINES: tuple[LineItem, ...] = (
    LineItem("GLFP-04821", "JIC 37 Male Straight Adapter, -6 (9/16-18 UNF), Steel ZN", 48, "EA", 3.08),
    LineItem("GLFP-04824", "JIC 37 Male Elbow 90 deg, -6 (9/16-18 UNF), Forged Steel", 25, "EA", 7.82),
    LineItem("GLFP-04831", "JIC Male x NPT Male Adapter, -6 x 1/4in NPT, Steel", 30, "EA", 3.75),
    LineItem("GLFP-04840", "SAE ORB Male Straight, -6 (9/16-18 UNF), Buna-N, Steel", 22, "EA", 5.02),
    LineItem("GLFP-05110", "Hydraulic Hose SAE 100R2AT, -6 (3/8 ID), 2-Wire", 175, "FT", 3.47),
    LineItem("GLFP-05115", "Hydraulic Hose SAE 100R2AT, -8 (1/2 ID), 2-Wire", 225, "FT", 4.58),
    LineItem("GLFP-05100", "Hydraulic Hose SAE 100R1AT, -4 (1/4 ID), 1-Wire", 125, "FT", 2.05),
    LineItem("GLFP-05220", "Hose End Fitting, Female JIC Swivel -8, Crimp 2-Wire", 60, "EA", 6.24),
    LineItem("GLFP-06010", "QD Coupler ISO 7241-A (AG), -6 Body x 3/8in NPT F, Steel", 6, "EA", 29.15),
    LineItem("GLFP-06050", "Inline Check Valve, 1/2in NPT F, 5 PSI Crack, Steel", 4, "EA", 25.80),
    LineItem("GLFP-11200", "QC Coupler Industrial M-Style, 1/4in NPT F, Steel", 18, "EA", 6.45),
    LineItem("GLFP-11210", "QC Plug Industrial M-Style, 1/4in NPT M, Hardened Steel", 24, "EA", 3.55),
)


def _thin_border() -> Border:
    side = Side(style="thin", color="808080")
    return Border(left=side, right=side, top=side, bottom=side)


def build_workbook() -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Weekly Reorder"

    # Column widths — tuned for a landscape-ish reorder sheet
    widths = {"A": 6, "B": 16, "C": 54, "D": 8, "E": 7, "F": 12, "G": 14}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # Row 1-3 — text letterhead (mid-market template, no embedded logo)
    ws["A1"] = "GREAT LAKES FLUID POWER GROUP"
    ws["A1"].font = Font(name="Arial", size=16, bold=True, color="1F3864")
    ws.merge_cells("A1:G1")
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")

    ws["A2"] = "1847 Leonard Street NW  •  Grand Rapids, MI 49504  •  (616) 458-7700"
    ws["A2"].font = Font(name="Arial", size=9, color="595959")
    ws.merge_cells("A2:G2")

    ws["A3"] = "Weekly Reorder — Hydraulics & Pneumatics"
    ws["A3"].font = Font(name="Arial", size=10, italic=True, color="595959")
    ws.merge_cells("A3:G3")

    # Row 5 — PURCHASE ORDER banner
    ws["A5"] = "PURCHASE ORDER"
    ws["A5"].font = Font(name="Arial", size=14, bold=True)
    ws.merge_cells("A5:G5")
    ws["A5"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[5].height = 22

    # Rows 7-11 — metadata block (two-column label:value pairs)
    meta_font_label = Font(name="Arial", size=10, bold=True)
    meta_font_val = Font(name="Arial", size=10)

    metadata = [
        ("PO Number:",         "PO-2026-05847",             "Buyer:",              "David Vance"),
        ("PO Date:",           "04/14/2026",                "Buyer Email:",        "dvance@glfp.com"),
        ("Required By:",       "04/21/2026",                "Buyer Phone:",        "(616) 458-7720"),
        ("Ship To:",           "GLFP-GRR-MAIN",             "Payment Terms:",      "Net 30"),
        ("Vendor:",            "Grafton-Reese MRO",         "Ship Via:",           "Vendor routing"),
    ]
    for i, (lab_l, val_l, lab_r, val_r) in enumerate(metadata, start=7):
        ws.cell(row=i, column=1, value=lab_l).font = meta_font_label
        ws.cell(row=i, column=2, value=val_l).font = meta_font_val
        ws.cell(row=i, column=5, value=lab_r).font = meta_font_label
        ws.cell(row=i, column=6, value=val_r).font = meta_font_val
        ws.cell(row=i, column=6).alignment = Alignment(horizontal="left")
        ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=4)
        ws.merge_cells(start_row=i, start_column=6, end_row=i, end_column=7)

    # Row 13 — column header row
    header_row = 13
    headers = ["Line", "GLFP Part #", "Description", "Qty", "UOM", "Unit Price", "Ext. Price"]
    header_fill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
    header_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    header_align = Alignment(horizontal="center", vertical="center")
    for col, text in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=text)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = _thin_border()
    ws.row_dimensions[header_row].height = 20

    # Line item rows
    body_font = Font(name="Arial", size=10)
    money_fmt = "$#,##0.00"
    qty_fmt = "#,##0"
    for i, line in enumerate(LINES, start=1):
        r = header_row + i
        ws.cell(row=r, column=1, value=i).alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=2, value=line.glfp_part)
        ws.cell(row=r, column=3, value=line.description)
        ws.cell(row=r, column=3).alignment = Alignment(wrap_text=True, vertical="center")
        ws.cell(row=r, column=4, value=line.qty).number_format = qty_fmt
        ws.cell(row=r, column=4).alignment = Alignment(horizontal="right")
        ws.cell(row=r, column=5, value=line.uom).alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=6, value=line.unit_price).number_format = money_fmt
        # Extended price as Excel formula (how a real template would calculate it)
        ws.cell(row=r, column=7, value=f"=D{r}*F{r}").number_format = money_fmt

        for col in range(1, 8):
            ws.cell(row=r, column=col).font = body_font
            ws.cell(row=r, column=col).border = _thin_border()

        ws.row_dimensions[r].height = 18

    last_line_row = header_row + len(LINES)

    # Subtotal + total block, two rows below table
    total_row = last_line_row + 2
    ws.cell(row=total_row, column=6, value="Subtotal").font = Font(name="Arial", size=10, bold=True)
    ws.cell(row=total_row, column=6).alignment = Alignment(horizontal="right")
    subtotal_formula = f"=SUM(G{header_row + 1}:G{last_line_row})"
    ws.cell(row=total_row, column=7, value=subtotal_formula).number_format = money_fmt
    ws.cell(row=total_row, column=7).font = Font(name="Arial", size=10, bold=True)
    ws.cell(row=total_row, column=7).border = _thin_border()

    ws.cell(row=total_row + 1, column=6, value="Tax (resale exempt)").font = Font(name="Arial", size=10)
    ws.cell(row=total_row + 1, column=6).alignment = Alignment(horizontal="right")
    ws.cell(row=total_row + 1, column=7, value=0.00).number_format = money_fmt
    ws.cell(row=total_row + 1, column=7).font = Font(name="Arial", size=10)

    ws.cell(row=total_row + 2, column=6, value="TOTAL").font = Font(name="Arial", size=11, bold=True)
    ws.cell(row=total_row + 2, column=6).alignment = Alignment(horizontal="right")
    ws.cell(row=total_row + 2, column=7, value=f"=G{total_row}+G{total_row + 1}").number_format = money_fmt
    ws.cell(row=total_row + 2, column=7).font = Font(name="Arial", size=11, bold=True)
    ws.cell(row=total_row + 2, column=7).border = Border(top=Side(style="thin"), bottom=Side(style="double"))

    # Special instructions footer
    notes_row = total_row + 5
    ws.cell(row=notes_row, column=1, value="Special Instructions:").font = Font(name="Arial", size=10, bold=True)
    ws.cell(row=notes_row + 1, column=1, value=(
        "Hose cut lengths acceptable. Consolidate to one shipment — do not split line 5/6/7 across "
        "multiple POs. If QD couplers (line 9) are backordered, notify D. Vance before substituting."
    ))
    ws.cell(row=notes_row + 1, column=1).alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=notes_row + 1, start_column=1, end_row=notes_row + 2, end_column=7)
    ws.row_dimensions[notes_row + 1].height = 32

    # Signature block
    sig_row = notes_row + 4
    ws.cell(row=sig_row, column=1, value="Authorized By: David Vance, Senior Buyer — Hydraulics")
    ws.cell(row=sig_row, column=1).font = Font(name="Arial", size=9, italic=True)
    ws.cell(row=sig_row + 1, column=1, value="Electronic signature on file. PO valid without wet signature.")
    ws.cell(row=sig_row + 1, column=1).font = Font(name="Arial", size=9, italic=True, color="595959")

    # Freeze header band so the line item rows scroll under the PO metadata
    ws.freeze_panes = "A14"

    return wb


def main() -> Path:
    out = Path(__file__).resolve().parents[2] / "data" / "excel" / "glfp_weekly_reorder_2026-04-14.xlsx"
    out.parent.mkdir(parents=True, exist_ok=True)
    wb = build_workbook()
    wb.save(out)
    return out


if __name__ == "__main__":
    path = main()
    print(f"wrote: {path}")
